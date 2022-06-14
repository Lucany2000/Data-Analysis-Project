import os
from turtle import color
from openpyxl import load_workbook, Workbook
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np

data = os.listdir("analysis_project") #the folder where all the tsv files are kept

def Excel(): #stores all of the tsv files into a single excel spreadsheet with each tsv file having its own sheet
    if os.path.exists('/analysis_project.xlsx') == False:
        wb = Workbook()
        del wb['Sheet'] 
    else:    
        wb = load_workbook('analysis_project.xlsx')

    writer = pd.ExcelWriter('analysis_project.xlsx', engine = 'openpyxl')
    writer.book = wb

    for d in data:
        fname = d.replace(".pooled.tsv","")
        if fname in wb.sheetnames:
            continue
        else:
            df = pd.read_csv(f'analysis_project\{d}', sep='\t')
            df.to_excel(writer, sheet_name=fname)

    wb.save('analysis_project.xlsx')
    writer.close()

#Excel()

tsv_library = {} #an empty libray for the dataframes
for d in data: #takes every tsv file from the folder
    fname = d.replace(".pooled.tsv","") #simplifies the tsv filenames
    #stores each dataframe in a dictionary Ex. HPAP001: the entire dataframe for that tsv file
    tsv_library[fname] = pd.read_csv(f'analysis_project\{d}', sep='\t') 

def copies_avg_v_identify_correlation():
    for tsv in tsv_library: #for every tsv file in the library
        #takes all of the values in the copies and avg_v_identity columns and calculates the correlation coefficient per tsv file 
        cor = np.corrcoef(tsv_library[tsv].avg_v_identity,tsv_library[tsv].copies)[0,1]
        #key: 
        #-1: complete negative correlation
        #1: complete positive correlation
        #0: no correlation
        print(f"{tsv}:",round(cor,2))
    figure, axis = plt.subplots(2,3) #creates subplots
    col = 0 #columns
    row = 0 #rows
    for tsv in tsv_library: 
        #creates scatter plot for each tsv file, each with a title, more accurate x/y axis, legend, and line of best fit 
        axis[row,col].scatter(tsv_library[tsv].avg_v_identity,tsv_library[tsv].copies)
        axis[row,col].plot(np.unique(tsv_library[tsv].avg_v_identity), np.poly1d(np.polyfit(tsv_library[tsv].avg_v_identity, tsv_library[tsv].avg_v_identity, 1))(np.unique(tsv_library[tsv].avg_v_identity)),color='k')
        axis[row,col].set_title(tsv)
        axis[row, col].ticklabel_format(useOffset=False)
        axis[row,col].legend(['Points','Line of Best Fit'])
        #shortcut for plotting all the subplots
        if col <= 1:
            col += 1
        else:
            row += 1
            col = 0
    #sets overall names for x and y axis        
    axis[1,0].set_xlabel('avg_v_identity') #labels the x axis
    axis[1,0].set_ylabel('copies') #labels the y axis
    plt.show() #displays subplots

#copies_avg_v_identify_correlation()

def V_gene_usage():
    Vg_copies = {} #empty dictionary for v_genes weighted by copies
    Vg_clones = {} #empty dictionary for v_genes weighted by donors
    for tsv in tsv_library: #takes every tsv file from the folder
        df = tsv_library[tsv] #shortcut to access the tsv file dataframe
        Vg = set(df.v_gene) #stores all v_gene names in a set: a list that doesn't contain any duplicates
        #gets the total number of copies per v_gene and stores in a dictionary with the v_gene as they key
        Vg_copies[tsv] = {copy : sum(df.loc[df['v_gene'] == copy]['copies'].values) for copy in Vg}
        #gets the total number of times that a v_gene appears within the donor with the v_gene as the key
        Vg_clones[tsv] = {clone : df['v_gene'].value_counts()[clone] for clone in Vg}
    for tsv in Vg_copies: #both Vg_clones and Vg_copies are always equal in contents and keys
        figure, (ax1, ax2) = plt.subplots(1,2) #creates a subplot for 2 pie charts
        figure.suptitle(tsv) #adds title to subplot
        #sorts their dictionaries by largest number
        cl = dict(sorted(Vg_clones[tsv].items(),key= lambda x:x[1]))
        co = dict(sorted(Vg_copies[tsv].items(),key= lambda x:x[1]))
        #combines any v_gene that takes up less than 1% of the total and stores it in "other"   
        temp = {x:co[x] for x in co if co[x]/sum(co.values())*100 < 1}
        for v in temp.keys():
            co.pop(v)
        co["other"] = round(sum(temp.values()))
        temp = {x:cl[x] for x in cl if cl[x]/sum(cl.values())*100 < 1}
        for v in temp.keys():
            cl.pop(v)
        cl["other"] = round(sum(temp.values()))
        #pie charts that show the how much of a percentage each v_gene takes up
        ax1.pie(cl.values(), labels=cl.keys(), autopct='%1.1f%%', textprops={'size': 'x-small'}, 
                wedgeprops={'linewidth': 3.0, 'edgecolor': 'white'})
        #sets name for pie charts        
        ax1.set_title("V-genes : clones")        
        ax2.pie(co.values(), labels=co.keys(), autopct='%1.1f%%', textprops={'size': 'x-small'}, 
                wedgeprops={'linewidth': 3.0, 'edgecolor': 'white'})
        ax2.set_title("V-genes : copies")
        #displays subplot that has pie charts                
        plt.show()

V_gene_usage()    
